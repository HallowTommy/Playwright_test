from __future__ import annotations
from pathlib import Path
import csv

def read_phones(path: Path) -> list[tuple[str,str]]:
    """
    returns list of (raw_phone, raw_phone) — нормализацию делаем в генераторе/или отдельном normalize.py
    """
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            rdr = csv.DictReader(f)
            if "phone" not in (rdr.fieldnames or []):
                raise ValueError("CSV должен иметь колонку phone")
            return [(row["phone"].strip(), row["phone"].strip()) for row in rdr if (row.get("phone") or "").strip()]

    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as e:
            raise RuntimeError("Нужен openpyxl для .xlsx: python -m pip install openpyxl") from e
        wb = load_workbook(path)
        ws = wb.active
        headers = [str(ws.cell(1,c).value or "").strip() for c in range(1, ws.max_column+1)]
        if "phone" not in headers:
            raise ValueError("XLSX должен иметь колонку phone в первой строке")
        col = headers.index("phone")+1
        out = []
        for r in range(2, ws.max_row+1):
            v = ws.cell(r,col).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append((s,s))
        wb.close()
        return out

    raise ValueError("Поддерживаются только .csv или .xlsx")
